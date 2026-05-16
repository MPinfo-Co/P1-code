"""/custom_table_data_input router — manual data maintenance for custom tables."""

import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.api.schema.fn_custom_table_data_input import (
    AuthorizedTableItem,
    AuthorizedTableListOut,
    ImportOut,
    RecordAddOut,
    RecordAddRequest,
    RecordDelOut,
    TableFieldItem,
    TableOptionItem,
    TableOptionsOut,
    TableRecordItem,
    TableRecordsData,
    TableRecordsOut,
)
from app.db.connector import get_db
from app.db.models.fn_custom_table import (
    CustomTable,
    CustomTableField,
    CustomTableRecord,
    RoleCustomTable,
)
from app.db.models.function_access import FunctionItems, RoleFunction
from app.db.models.user_role import UserRole
from app.logger_utils import get_system_logger
from app.utils.util_store import AuthContext, authenticate

router = APIRouter(prefix="/custom_table_data_input", tags=["custom_table_data_input"])
system_logger = get_system_logger()

FN_CUSTOM_TABLE_DATA_INPUT = "fn_custom_table_data_input"


def _is_numeric(value: object) -> bool:
    """Return True if value can be converted to float."""
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value))
        return True
    except ValueError:
        return False


def _has_fn_permission(user_id: int, db: Session) -> bool:
    """Return True if the user has fn_custom_table_data_input function permission."""
    fn = (
        db.query(FunctionItems)
        .filter(FunctionItems.function_code == FN_CUSTOM_TABLE_DATA_INPUT)
        .first()
    )
    if fn is None:
        return False
    return (
        db.query(RoleFunction)
        .join(UserRole, RoleFunction.role_id == UserRole.role_id)
        .filter(
            UserRole.user_id == user_id,
            RoleFunction.function_id == fn.function_id,
        )
        .first()
        is not None
    )


def _get_authorized_table_ids(user_id: int, db: Session) -> set[int]:
    """Return set of table_ids authorized for the user's roles."""
    role_ids = [
        r.role_id for r in db.query(UserRole).filter(UserRole.user_id == user_id).all()
    ]
    if not role_ids:
        return set()
    rows = db.query(RoleCustomTable).filter(RoleCustomTable.role_id.in_(role_ids)).all()
    return {r.table_id for r in rows}


def _check_fn_permission(user_id: int, db: Session) -> None:
    """Raise 403 if user does not have fn_custom_table_data_input permission."""
    if not _has_fn_permission(user_id, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您沒有執行此操作的權限",
        )


def _check_table_authorized(user_id: int, table_id: int, db: Session) -> None:
    """Raise 403 if the target table is not authorized for the user's roles."""
    authorized = _get_authorized_table_ids(user_id, db)
    if table_id not in authorized:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="您沒有執行此操作的權限",
        )


def _check_table_exists(table_id: int, db: Session) -> CustomTable:
    """Raise 404 if the table does not exist; otherwise return the table."""
    table = db.query(CustomTable).filter(CustomTable.id == table_id).first()
    if table is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="資料表不存在",
        )
    return table


@router.get("/tables", response_model=AuthorizedTableListOut)
def list_authorized_tables(
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
) -> AuthorizedTableListOut:
    """Return custom tables authorized for the current user. Requires fn_custom_table_data_input permission."""
    _check_fn_permission(auth.user_id, db)

    authorized_ids = _get_authorized_table_ids(auth.user_id, db)
    if not authorized_ids:
        return AuthorizedTableListOut(data=[])

    tables = db.query(CustomTable).filter(CustomTable.id.in_(authorized_ids)).all()
    data = [
        AuthorizedTableItem(id=t.id, name=t.name, description=t.description)
        for t in tables
    ]
    return AuthorizedTableListOut(data=data)


@router.get("/options", response_model=TableOptionsOut)
def get_table_options(
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
) -> TableOptionsOut:
    """Return all custom tables as [{ id, name }] for role dialog checkboxes."""
    tables = db.query(CustomTable).order_by(CustomTable.name.asc()).all()
    data = [TableOptionItem(id=t.id, name=t.name) for t in tables]
    return TableOptionsOut(data=data)


@router.get("/tables/{id}/records", response_model=TableRecordsOut)
def list_table_records(
    id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
) -> TableRecordsOut:
    """Return all records for a custom table along with field definitions."""
    _check_fn_permission(auth.user_id, db)
    _check_table_exists(id, db)
    _check_table_authorized(auth.user_id, id, db)

    fields = (
        db.query(CustomTableField)
        .filter(CustomTableField.table_id == id)
        .order_by(CustomTableField.sort_order.asc())
        .all()
    )
    records = (
        db.query(CustomTableRecord)
        .filter(CustomTableRecord.table_id == id)
        .order_by(CustomTableRecord.updated_at.desc())
        .all()
    )

    field_items = [
        TableFieldItem(
            id=f.id,
            field_name=f.field_name,
            field_type=f.field_type,
            sort_order=f.sort_order,
        )
        for f in fields
    ]
    record_items = [
        TableRecordItem(id=r.id, data=r.data, updated_at=r.updated_at) for r in records
    ]
    return TableRecordsOut(
        data=TableRecordsData(fields=field_items, records=record_items)
    )


@router.post(
    "/tables/{id}/records",
    response_model=RecordAddOut,
    status_code=status.HTTP_201_CREATED,
)
def add_table_record(
    id: int,
    payload: RecordAddRequest,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
) -> RecordAddOut:
    """Add a single record to a custom table."""
    _check_fn_permission(auth.user_id, db)
    _check_table_exists(id, db)
    _check_table_authorized(auth.user_id, id, db)

    # Validate field types
    fields = db.query(CustomTableField).filter(CustomTableField.table_id == id).all()
    field_type_map = {f.field_name: f.field_type for f in fields}
    for field_name, value in payload.data.items():
        if field_type_map.get(field_name) == "number":
            if not _is_numeric(value):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"欄位「{field_name}」須為數值",
                )

    record = CustomTableRecord(
        table_id=id,
        data=payload.data,
        updated_by=auth.user_id,
    )
    db.add(record)
    db.commit()
    system_logger.info(f"User {auth.user_id} added record to custom table {id}")
    return RecordAddOut()


@router.delete("/tables/{id}/records/{record_id}", response_model=RecordDelOut)
def delete_table_record(
    id: int,
    record_id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
) -> RecordDelOut:
    """Delete a single record from a custom table."""
    _check_fn_permission(auth.user_id, db)
    _check_table_exists(id, db)
    _check_table_authorized(auth.user_id, id, db)

    record = (
        db.query(CustomTableRecord)
        .filter(
            CustomTableRecord.id == record_id,
            CustomTableRecord.table_id == id,
        )
        .first()
    )
    if record is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="記錄不存在",
        )

    db.delete(record)
    db.commit()
    system_logger.info(
        f"User {auth.user_id} deleted record {record_id} from custom table {id}"
    )
    return RecordDelOut()


@router.post("/tables/{id}/import", response_model=ImportOut)
def import_table_records(
    id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
) -> ImportOut:
    """Import records from an .xlsx file (all-or-nothing)."""
    _check_fn_permission(auth.user_id, db)
    _check_table_exists(id, db)
    _check_table_authorized(auth.user_id, id, db)

    # Validate file extension
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="僅接受 .xlsx 格式檔案",
        )

    # Load field definitions
    fields = (
        db.query(CustomTableField)
        .filter(CustomTableField.table_id == id)
        .order_by(CustomTableField.sort_order.asc())
        .all()
    )
    field_type_map = {f.field_name: f.field_type for f in fields}

    # Parse xlsx
    content = file.file.read()
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ImportOut()

    header_row = rows[0]
    headers = [str(h) if h is not None else "" for h in header_row]
    data_rows = rows[1:]

    # Validate all data rows before writing
    error_rows: list[int] = []
    for row_idx, row in enumerate(data_rows, start=2):  # row 2 is first data row
        row_data = dict(zip(headers, row))
        for field_name, value in row_data.items():
            if field_type_map.get(field_name) == "number" and value is not None:
                if not _is_numeric(value):
                    error_rows.append(row_idx)
                    break

    if error_rows:
        row_list = "、".join(str(r) for r in error_rows)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"第 {row_list} 列格式有誤，所有資料均未寫入",
        )

    # Write all rows
    for row in data_rows:
        row_data = dict(zip(headers, row))
        # Convert number fields
        normalized: dict = {}
        for k, v in row_data.items():
            if field_type_map.get(k) == "number" and v is not None:
                normalized[k] = float(v)
            else:
                normalized[k] = v
        record = CustomTableRecord(
            table_id=id,
            data=normalized,
            updated_by=auth.user_id,
        )
        db.add(record)

    db.commit()
    system_logger.info(
        f"User {auth.user_id} imported {len(data_rows)} records to custom table {id}"
    )
    return ImportOut()


@router.get("/tables/{id}/format")
def download_table_format(
    id: int,
    db: Session = Depends(get_db),
    auth: AuthContext = Depends(authenticate),
):
    """Download a blank .xlsx template with header row based on field definitions."""
    _check_fn_permission(auth.user_id, db)
    _check_table_exists(id, db)
    _check_table_authorized(auth.user_id, id, db)

    fields = (
        db.query(CustomTableField)
        .filter(CustomTableField.table_id == id)
        .order_by(CustomTableField.sort_order.asc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.append([f.field_name for f in fields])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="format.xlsx"'},
    )
